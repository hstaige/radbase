import os

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

pd.set_option('display.width', 200)
pd.set_option('display.max_columns', None)

def compare_rel_results(filename, rel1, rel2, lsuffix='x', rsuffix='y'):
    rel1, rel2 = rel1.copy(), rel2.copy()
    rels = [rel1, rel2]
    for i, rel in enumerate(rels):
        if ('Z1' not in rel.columns or 'A1' not in rel.columns):
            raise ValueError(f'one of the dfs is missing \'A1\' or \'Z1\' or \'iso_idx1\'')
        if ('Z2' not in rel.columns or 'A2' not in rel.columns):
            raise ValueError(f'one of the dfs is missing \'A2\' or \'Z2\' or \'iso_idx2\'')

        rel = rel.astype({'Z1': int, 'A1': int, 'Z2': int, 'A2': int})
        rels[i] = rel

    merge_df = pd.merge(*rels, on=['Z1', 'A1', 'Z2', 'A2'], suffixes=('_' + lsuffix, '_' + rsuffix))
    merge_df['diff_dR'] = merge_df[f'Unc_{lsuffix}'] - merge_df[f'Unc_{rsuffix}']
    merge_df['diff_DdR'] = merge_df[f'DdR_{lsuffix}'] - merge_df[f'DdR_{rsuffix}']

    columns = ['Z1', 'A1', 'Z2', 'A2', f'Unc_{lsuffix}', f'DdR_{lsuffix}', f'Unc_{rsuffix}', f'DdR_{rsuffix}',
               'diff_dR', 'diff_DdR']
    merge_df = merge_df[columns]
    merge_df = merge_df.round({c: 4 for c in columns[4:]})

    # note: xlsxwriter overwrites old sheets, need to read them in first if file exists already
    # if os.path.exists(filename):
    #     workbook = openpyxl.load_workbook(filename)
    #     writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    #     writer.book = workbook
    #     writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    # else:
    #     writer = pd.ExcelWriter(filename, engine='xlsxwriter')

    mode = 'a' if os.path.exists(filename) and False else 'w'
    with pd.ExcelWriter(filename, engine='openpyxl', if_sheet_exists="error", mode=mode) as writer:
        merge_df.to_excel(writer, sheet_name='relative', index=False)
        worksheet = writer.sheets['relative']
        for col in worksheet.columns:  # loop through all columns
            series = merge_df.iloc[:, col[0].column - 1]
            max_len = max((
                series.astype(str).map(len).max(),  # len of largest item
                len(str(series.label))  # len of column name/header
            )) + 1  # adding a little extra space)
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)].width = max_len  # set column width

    return merge_df


def compare_abs_results(filename, absl1, absl2, lsuffix='x', rsuffix='y'):
    absl1, absl2 = absl1.copy(), absl2.copy()
    absls = [absl1, absl2]
    for i, absl in enumerate(absls):
        print(absl.columns)
        if ('Z' not in absl.columns or 'A' not in absl.columns) and 'iso_idx' not in absl.columns:
            raise ValueError(f'one of the dfs is missing \'A1\' or \'Z1\' or \'iso_idx1\'')

        if 'Z' not in absl.columns or 'A' not in absl.columns:
            merge_df = pd.merge(absl, iso, left_on='iso_idx', right_on='idx')
            absl = merge_df.drop(columns=['iso_idx', 'idx'])

        absl = absl.astype({'Z': int, 'A': int})
        absls[i] = absl

    merge_df = pd.merge(*absls, on=['Z', 'A'], suffixes=('_' + lsuffix, '_' + rsuffix))
    merge_df['diff_R'] = merge_df[f'Value_{lsuffix}'] - merge_df[f'Value_{rsuffix}']
    merge_df['diff_DR'] = merge_df[f'Unc_{lsuffix}'] - merge_df[f'Unc_{rsuffix}']

    columns = ['Z', 'A', f'Value_{lsuffix}', f'Unc_{lsuffix}', f'Value_{rsuffix}', f'Unc_{rsuffix}',
               'diff_R', 'diff_DR']
    merge_df = merge_df[columns]
    merge_df = merge_df.astype({'Z': int, 'A': int})
    merge_df = merge_df.round({c: 4 for c in columns[2:]})

    # if os.path.exists(filename) :
    #     workbook = openpyxl.load_workbook(filename)
    #     writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    #     writer.book = workbook
    #     writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    # else:
    #     writer = pd.ExcelWriter(filename, engine='xlsxwriter')

    mode = 'a' if os.path.exists(filename) else 'w'
    with pd.ExcelWriter(filename, engine='openpyxl', mode=mode) as writer:
        print(writer.sheets)
        merge_df.to_excel(writer, sheet_name='absolute', index=False)
        worksheet = writer.sheets['absolute']
        for col in worksheet.columns:  # loop through all columns
            series = merge_df.iloc[:, col[0].column - 1]
            max_len = max((
                series.astype(str).map(len).max(),  # len of largest item
                len(str(series.label))  # len of column name/header
            )) + 1  # adding a little extra space)
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)].width = max_len  # set column width

    return merge_df


abs_df = pd.read_excel('./NuclearRadius_all_sheets_renamed.xlsx', skiprows=5, usecols=[1, 2, 5, 6, 7, 8])
abs_df = abs_df.dropna(subset=['dRav'], how='any')

abs_df = abs_df[~abs_df['Z'].isin([0, 1])]
abs_df.rename(columns={'Rav': 'Value', 'dRav': 'Unc'}, inplace=True)
abs_df.reset_index(drop=True, inplace=True)

compare_abs_results('comparisons/angeli_abs_abs.xlsx', abs_df, pd.read_csv(
    '../outputs/absonly/absolute_radii_absonly.csv'),
                    lsuffix='ar', rsuffix='hr')

# rel_noniso_df = pd.read_excel('./NuclearRadius_all_sheets_renamed.xlsx', usecols=[0, 2, 3, 5, 24, 28, 18],
#                               sheet_name='dRem_non_iso_11')
# rel_noniso_df.drop(rel_noniso_df.tail(2).index, inplace=True)
# rel_noniso_df.dropna(how='any', subset=['d0'], inplace=True)
# rel_noniso_df.ffill(inplace=True)
# rel_noniso_df.rename(columns={'d0': 'Unc_', 'Dd0': 'DdR'}, inplace=True)
# rel_noniso_df.reset_index(drop=True, inplace=True)
#
# compare_rel_results('../comparisons/angeli_dnoniso_noniso_relonly.xlsx', rel_noniso_df,
#                     pd.read_csv('outputs/noniso_relonly/relative_radii_noniso_relonly.csv'), iso_df,
#                     lsuffix='adremnoniso', rsuffix='hdremnoniso')
#
# rel_noniso_df = pd.read_excel('./NuclearRadius_all_sheets_renamed.xlsx', usecols=[7, 9, 10, 12, 48, 49,50, 18],
#                               sheet_name='Rem_non_iso_11')
# rel_noniso_df = rel_noniso_df[rel_noniso_df['Include'] != 'No']
# rel_noniso_df.drop(rel_noniso_df.tail(2).index, inplace=True)
# rel_noniso_df.dropna(how='any', subset=['Unc_+'], inplace=True)
# rel_noniso_df.ffill(inplace=True)
# rel_noniso_df.rename(columns={'Unc_+': 'Unc_', 'DdR+': 'DdR'}, inplace=True)
# rel_noniso_df.reset_index(drop=True, inplace=True)
#
# compare_rel_results('../comparisons/angeli_noniso_noniso.xlsx', rel_noniso_df,
#                     pd.read_csv('outputs/noniso/relative_radii_noniso.csv'), iso_df,
#                     lsuffix='aremnoniso', rsuffix='hremnoniso')
#
# absl_noniso_df = pd.read_excel('./NuclearRadius_all_sheets_renamed.xlsx', usecols=[0, 2, 46, 47, 18, 50],
#                                sheet_name='Rem_non_iso_11')
# absl_noniso_df.drop(absl_noniso_df.tail(2).index, inplace=True)
# absl_noniso_df.dropna(how='any', subset=['R+'], inplace=True)
#
# cols = ['Z', 'A']
# absl_noniso_df[cols] = absl_noniso_df[cols].ffill()
#
# absl_noniso_df = absl_noniso_df[absl_noniso_df['Include'] != 'No']
# absl_noniso_df.rename(columns={'R+': 'R', 'Unc_+': 'Unc_'}, inplace=True)
# absl_noniso_df.reset_index(drop=True, inplace=True)
#
# # print(absl_noniso_df, pd.read_csv('noniso/absolute_radii_noniso.csv'))
# # print(pd.merge(absl_noniso_df, pd.read_csv('noniso/absolute_radii_noniso.csv'), on=['Z', 'A']).head(20))
#
# compare_abs_results('../comparisons/angeli_noniso_noniso.xlsx', absl_noniso_df,
#                     pd.read_csv('outputs/noniso/absolute_radii_noniso.csv'), iso_df,
#                     lsuffix='aremnoniso', rsuffix='hremnoniso')
